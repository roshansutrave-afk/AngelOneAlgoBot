"""
tests/test_market_charges_excel.py
Market hours, charges, and excel logger tests — no broker connection.
"""
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pytest

from core.charges import ProductType, calculate_charges, round_trip_charges
from core.excel_logger import ExcelLogger
from core.market_hours import is_trading_day, NSE_HOLIDAYS
from core.strategy import Action


# ── Charges ───────────────────────────────────────────────────────────────────

def test_intraday_stt_sell_side_only():
    buy  = calculate_charges(500, 100, ProductType.INTRADAY, "BUY")
    sell = calculate_charges(500, 100, ProductType.INTRADAY, "SELL")
    assert buy.stt == 0.0
    assert sell.stt > 0.0
    assert abs(sell.stt - 500 * 100 * 0.00025) < 0.001


def test_delivery_stt_both_sides():
    buy  = calculate_charges(500, 100, ProductType.DELIVERY, "BUY")
    sell = calculate_charges(500, 100, ProductType.DELIVERY, "SELL")
    assert buy.stt > 0 and sell.stt > 0


def test_stamp_buy_only():
    buy  = calculate_charges(200, 50, ProductType.INTRADAY, "BUY")
    sell = calculate_charges(200, 50, ProductType.INTRADAY, "SELL")
    assert buy.stamp_duty > 0
    assert sell.stamp_duty == 0.0


def test_gst_on_brokerage_and_txn():
    c = calculate_charges(1000, 10, ProductType.INTRADAY, "BUY", 20)
    assert abs(c.gst - (c.brokerage + c.exchange_txn) * 0.18) < 0.01


def test_round_trip_sanity():
    rt = round_trip_charges(500, 510, 100, ProductType.INTRADAY, 20)
    assert rt.total >= 40           # at least 2× ₹20 brokerage
    assert rt.total / (500 * 100) < 0.02   # less than 2% of turnover


def test_fo_futures_lower_txn():
    eq  = calculate_charges(500, 100, ProductType.INTRADAY,   "SELL")
    fut = calculate_charges(500, 100, ProductType.FO_FUTURES,  "SELL")
    assert fut.exchange_txn < eq.exchange_txn


def test_currency_zero_stt():
    c = calculate_charges(83.5, 1000, ProductType.CURRENCY, "SELL")
    assert c.stt == 0.0


# ── Market hours ──────────────────────────────────────────────────────────────

def test_saturday_not_trading():
    assert not is_trading_day(date(2025, 6, 21))


def test_sunday_not_trading():
    assert not is_trading_day(date(2025, 6, 22))


def test_nse_holidays_not_trading():
    for h in list(NSE_HOLIDAYS)[:3]:
        assert not is_trading_day(h)


def test_regular_weekday_is_trading():
    assert is_trading_day(date(2025, 6, 23))


# ── Excel logger ──────────────────────────────────────────────────────────────

def test_excel_entry_exit_round_trip():
    tmp = Path(tempfile.mkdtemp())
    xl  = ExcelLogger(data_dir=tmp, product_type=ProductType.INTRADAY, brokerage_flat=20)
    path = xl.open_session(datetime(2025, 6, 23, 9, 30))

    xl.log_entry("SBIN-EQ", "NSE", "Trend", "BUY", 10, 820.0,
                  datetime(2025, 6, 23, 9, 30, 0))
    xl.log_exit("SBIN-EQ", 835.0, datetime(2025, 6, 23, 11, 45, 0))
    xl.close()

    wb = openpyxl.load_workbook(path)
    ws = wb["Trade Log"]
    row = [ws.cell(row=3, column=c).value for c in range(1, 20)]

    assert row[2] == "SBIN-EQ"
    assert row[5] == "BUY"
    assert row[6] == 10
    assert row[7] == 820.0
    assert row[9] == 835.0
    assert row[11] == 135          # hold = 135 min
    assert row[12] > 0             # brokerage non-zero
    assert row[13] > 0             # STT non-zero
    assert row[15] == pytest.approx(150.0, abs=0.1)   # gross PnL = 15 × 10


def test_open_trade_written_immediately():
    tmp = Path(tempfile.mkdtemp())
    xl  = ExcelLogger(data_dir=tmp, product_type=ProductType.INTRADAY, brokerage_flat=20)
    path = xl.open_session(datetime(2025, 6, 23, 10, 0))
    xl.log_entry("INFY-EQ", "NSE", "Trend", "BUY", 5, 1540.0,
                  datetime(2025, 6, 23, 10, 0, 0))

    wb = openpyxl.load_workbook(path)
    ws = wb["Trade Log"]
    assert ws.cell(row=3, column=3).value == "INFY-EQ"
    assert ws.cell(row=3, column=10).value in ("", None)   # exit ₹ blank


def test_unknown_exit_does_not_raise():
    tmp = Path(tempfile.mkdtemp())
    xl  = ExcelLogger(data_dir=tmp, product_type=ProductType.INTRADAY, brokerage_flat=20)
    xl.open_session()
    xl.log_exit("UNKNOWN", 100.0)   # no matching open trade — must not raise
