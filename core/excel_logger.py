"""
core/excel_logger.py
Excel trade logger matching the Algo_Trading_Tracker format.

Trade Log columns:
  #, Date, Symbol, Exchange, Strategy, Side, Qty, Entry ₹, Entry Time,
  Exit ₹, Exit Time, Hold(min), Broker ₹, STT ₹, Other ₹,
  Gross P&L ₹, Net P&L ₹, Return %, Notes

Entry and exit are separate fill events. On BUY the row is written
immediately (amber — open). On the matching SELL the row is updated
in-place with exit data, charges, and P&L (green/red). File is
flushed to disk after every write.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XlSide
from openpyxl.utils import get_column_letter

from core.charges import ProductType, round_trip_charges

HEADER_BG  = "1F2D3D"
PROFIT_FG  = "00B050"
LOSS_FG    = "FF0000"
OPEN_FG    = "FF8C00"
BLUE_INPUT = "0000FF"
ALT_ROW_BG = "F2F2F2"

HEADERS = [
    "#", "Date", "Symbol", "Exchange", "Strategy", "Side", "Qty",
    "Entry ₹", "Entry Time", "Exit ₹", "Exit Time", "Hold(min)",
    "Broker ₹", "STT ₹", "Other ₹", "Gross P&L ₹", "Net P&L ₹",
    "Return %", "Notes",
]
COL_WIDTHS = [5, 12, 14, 10, 14, 6, 7, 10, 12, 10, 12, 10, 10, 10, 10, 12, 12, 10, 30]


class OpenTrade:
    __slots__ = ("trade_num", "row_idx", "symbol", "exchange",
                 "strategy", "side", "qty", "entry_price", "entry_time")

    def __init__(self, trade_num, row_idx, symbol, exchange,
                 strategy, side, qty, entry_price, entry_time):
        self.trade_num   = trade_num
        self.row_idx     = row_idx
        self.symbol      = symbol
        self.exchange    = exchange
        self.strategy    = strategy
        self.side        = side
        self.qty         = qty
        self.entry_price = entry_price
        self.entry_time  = entry_time


class ExcelLogger:
    """One instance per trading session. Call open_session() first."""

    def __init__(
        self,
        data_dir: Path,
        product_type: ProductType = ProductType.INTRADAY,
        brokerage_flat: float = 20.0,
    ):
        self.data_dir       = data_dir
        self.product_type   = product_type
        self.brokerage_flat = brokerage_flat

        self._wb: Optional[Workbook] = None
        self._ws  = None
        self._path: Optional[Path]   = None
        self._open_trades: dict[str, OpenTrade] = {}
        self._trade_count: int = 0
        self._next_row: int    = 3          # row 1 = title, row 2 = headers

    def open_session(self, date: Optional[datetime] = None) -> Path:
        date = date or datetime.now()
        self._path = self.data_dir / f"AlgoBot_TradeLog_{date.strftime('%Y%m%d')}.xlsx"
        self.data_dir.mkdir(parents=True, exist_ok=True)

        if self._path.exists():
            self._wb = openpyxl.load_workbook(self._path)
            self._ws = self._wb["Trade Log"]
            self._trade_count = self._ws.max_row - 2
            self._next_row    = self._ws.max_row + 1
        else:
            self._wb = Workbook()
            self._ws = self._wb.active
            self._ws.title = "Trade Log"
            self._write_header(date)
            self._next_row = 3

        return self._path

    # ── Public API ────────────────────────────────────────────────────────────

    def log_entry(
        self,
        symbol: str,
        exchange: str,
        strategy: str,
        side: str,
        qty: int,
        price: float,
        timestamp: Optional[datetime] = None,
        notes: str = "",
    ) -> None:
        if self._ws is None:
            raise RuntimeError("Call open_session() before logging.")
        ts = timestamp or datetime.now()
        self._trade_count += 1
        row = self._next_row
        self._next_row += 1

        self._open_trades[symbol] = OpenTrade(
            self._trade_count, row, symbol, exchange,
            strategy, side, qty, price, ts,
        )
        self._write_row(row, [
            self._trade_count, ts.strftime("%d-%b-%y"), symbol, exchange,
            strategy, side, qty, round(price, 2), ts.strftime("%H:%M:%S"),
            "", "", "", "", "", "", "", "", "", notes,
        ], status="open")
        self._flush()

    def log_exit(
        self,
        symbol: str,
        exit_price: float,
        timestamp: Optional[datetime] = None,
        notes: str = "",
    ) -> None:
        if self._ws is None:
            raise RuntimeError("Call open_session() before logging.")
        ts = timestamp or datetime.now()
        ot = self._open_trades.pop(symbol, None)

        if ot is None:
            self._trade_count += 1
            self._write_row(self._next_row, [
                self._trade_count, ts.strftime("%d-%b-%y"), symbol,
                "", "", "SELL", 0, "", "",
                round(exit_price, 2), ts.strftime("%H:%M:%S"),
                "", "", "", "", "", "", "", notes,
            ], status="closed")
            self._next_row += 1
            self._flush()
            return

        hold_mins = int((ts - ot.entry_time).total_seconds() / 60)
        costs     = round_trip_charges(ot.entry_price, exit_price, ot.qty,
                                        self.product_type, self.brokerage_flat)
        direction = 1 if ot.side == "BUY" else -1
        gross     = direction * (exit_price - ot.entry_price) * ot.qty
        net       = gross - costs.total
        ret_pct   = (net / (ot.entry_price * ot.qty) * 100) if ot.entry_price * ot.qty else 0.0

        self._write_row(ot.row_idx, [
            ot.trade_num, ot.entry_time.strftime("%d-%b-%y"), symbol,
            ot.exchange, ot.strategy, ot.side, ot.qty,
            round(ot.entry_price, 2), ot.entry_time.strftime("%H:%M:%S"),
            round(exit_price, 2), ts.strftime("%H:%M:%S"), hold_mins,
            round(costs.brokerage, 2), round(costs.stt, 2), round(costs.other, 2),
            round(gross, 2), round(net, 2), round(ret_pct, 2),
            notes or ot.strategy,
        ], status="profit" if net >= 0 else "loss")
        self._flush()

    def add_totals_row(self) -> None:
        if self._ws is None or self._next_row <= 3:
            return
        end = self._next_row - 1
        row = self._next_row
        for ci in range(1, len(HEADERS) + 1):
            cell = self._ws.cell(row=row, column=ci)
            if ci == 1:
                cell.value = "TOTALS"
            elif ci in (7, 13, 14, 15, 16, 17):
                letter = get_column_letter(ci)
                cell.value = f"=SUM({letter}3:{letter}{end})"
            cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color=HEADER_BG, fgColor=HEADER_BG)
        self._next_row += 1
        self._flush()

    def close(self) -> None:
        self.add_totals_row()
        self._flush()

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _write_header(self, date: datetime) -> None:
        ws = self._ws
        ws.merge_cells("A1:S1")
        c = ws["A1"]
        c.value = (f"TRADE LOG — AlgoBot Session  |  {date.strftime('%d %b %Y')}"
                   f"  |  NSE / BSE  |  AngelOne")
        c.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color=HEADER_BG, fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            cell.fill      = PatternFill("solid", start_color=HEADER_BG, fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = _thin()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 22
        ws.freeze_panes = "A3"

    def _write_row(self, row_idx: int, values: list, status: str) -> None:
        bg = ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(values, 1):
            cell = self._ws.cell(row=row_idx, column=ci, value=val)
            cell.border    = _thin()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill      = PatternFill("solid", start_color=bg, fgColor=bg)

            is_pnl = ci in (16, 17)
            is_ret = ci == 18
            if status == "open":
                fc = OPEN_FG
            elif status in ("profit", "loss") and (is_pnl or is_ret):
                fc = PROFIT_FG if status == "profit" else LOSS_FG
            elif ci in (8, 10):
                fc = BLUE_INPUT
            else:
                fc = "000000"

            cell.font = Font(name="Calibri", size=10, color=fc,
                              bold=ci in (1, 6, 16, 17))

            if ci in (8, 10):
                cell.number_format = "#,##0.00"
            elif ci in (13, 14, 15, 16, 17):
                cell.number_format = "#,##0.00;[Red](#,##0.00)"
            elif ci == 18:
                cell.number_format = "0.00%"
                if isinstance(val, (int, float)):
                    cell.value = val / 100

    def _flush(self) -> None:
        if self._wb and self._path:
            self._wb.save(self._path)


def _thin() -> Border:
    s = XlSide(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
